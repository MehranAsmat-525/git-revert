import pandas as pd
import os
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# --- Step 1: Load the latest Excel file ---
folder_path = 'D:/Main'
files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
if not files:
    print("No Excel files found.")
    exit()

full_paths = [os.path.join(folder_path, f) for f in files]
latest_file = max(full_paths, key=os.path.getmtime)

# --- Step 2: Read Excel data ---
df = pd.read_excel(latest_file)

# --- Step 3: Clean weight column ---
if 'Name' not in df.columns or 'Weight' not in df.columns:
    print("Required columns not found.")
    exit()

# Remove 'kg' or other non-numeric from weight
df['Weight'] = df['Weight'].astype(str).str.extract(r'(\d+)').astype(float)

# --- Step 4: Create and save the graph ---
plt.figure(figsize=(10, 6))
plt.bar(df['Name'], df['Weight'], color='skyblue')
plt.xlabel('Name')
plt.ylabel('Weight')
plt.title('Name vs Weight')
plt.xticks(rotation=45)
plt.tight_layout()

# Save the chart image
chart_path = os.path.join(folder_path, 'temp_chart.png')
plt.savefig(chart_path)
plt.close()

# --- Step 5: Insert graph into Excel as new sheet ---
wb = load_workbook(latest_file)
ws = wb.create_sheet(title="Graph")

img = ExcelImage(chart_path)
img.anchor = 'A1'  # Position where to place the image
ws.add_image(img)

# Save Excel file
wb.save(latest_file)

# --- Step 6: Cleanup temp image ---
os.remove(chart_path)

print(f"✅ Chart added to new sheet 'Graph' in {os.path.basename(latest_file)}")
